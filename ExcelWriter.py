import pandas as pd
from openpyxl import load_workbook

from constants.main_settings import APP_VERSION

class ExcelWriter:
    def __init__(self):

        self.SUMMARY_RESULTS_EXCEL_PATH = "wyniki/wyniki.xlsx"
        self.DATA_EXCEL_PATH = "wyniki/dane.xlsx"

        self.SUMMARY_RESULTS_ROW_STEP = 1
        self.DATA_ROW_STEP = 4

        self.row_summary_results = self._get_empty_cell_index(self.SUMMARY_RESULTS_EXCEL_PATH, "Sheet1", self.SUMMARY_RESULTS_ROW_STEP)
        self.row_data = self._get_empty_cell_index(self.DATA_EXCEL_PATH, "Sheet1", self.DATA_ROW_STEP)

        #self.are_correct_answers_saved = False

    def save_data_to_excel(self, correct_answers, human_answers, reaction_time, person):
        self.save_correct_answers(correct_answers)

        correct_answers_count = self._count_correct_answers(correct_answers, human_answers)
        df_results = pd.DataFrame(
            [APP_VERSION, person.ID, person.age, person.sex, correct_answers_count, len(human_answers)])
        with pd.ExcelWriter(self.SUMMARY_RESULTS_EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_results.T.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=self.row_summary_results, header=False,
                                  index=False)

        df_data = pd.DataFrame([person.ID], ["ID"])
        with pd.ExcelWriter(self.DATA_EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_data.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=self.row_data, header=False,
                             index=True)
        df_data = pd.DataFrame([human_answers, reaction_time], ["Odpowiedzi", "Czas reakcji [s]"])
        with pd.ExcelWriter(self.DATA_EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_data.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=self.row_data + 1, header=False,
                             index=True)

        self._move_rows_index()

    def _move_rows_index(self):
        self.row_summary_results += self.SUMMARY_RESULTS_ROW_STEP
        self.row_data += self.DATA_ROW_STEP

    def _count_correct_answers(self, a, b):
        """
        Count how many items are the same and in the same positions in two lists.

        :param a: First list
        :param b: Second list
        :return: Number of items that are the same and in the same position
        """
        return sum(1 for x, y in zip(a, b) if x == y)

    def _get_empty_cell_index(self, file_path, sheet_name, row_step = 4, start_row=2, column="B"):

        # Load the workbook and select the sheet
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]

        row = start_row

        while True:
            cell_value = sheet[f"{column}{row}"].value  # Get cell value
            if cell_value is None:
                return row - 1 # -1 because pd start saving from row below
            row += row_step  # Jump X rows ahead

    def save_correct_answers(self, correct_answers):
        df_results = pd.DataFrame(
            correct_answers)
        with pd.ExcelWriter(self.SUMMARY_RESULTS_EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_results.T.to_excel(writer, sheet_name='Poprawne_odpowiedzi', startcol=0, startrow=self.row_summary_results, header=False,
                                  index=False)

        #self.are_correct_answers_saved = True

    def get_last_person_ID(self):
        empty_cell_index = self._get_empty_cell_index(self.SUMMARY_RESULTS_EXCEL_PATH, "Sheet1", self.SUMMARY_RESULTS_ROW_STEP)
        print("Empty cell index: " +str(empty_cell_index))
        if empty_cell_index == 1:
            return 1
        else:
            wb = load_workbook(self.SUMMARY_RESULTS_EXCEL_PATH)
            sheet = wb["Sheet1"]
            row = empty_cell_index
            column = "B"
            last_ID = sheet[f"{column}{row}"].value  # Get cell value
            return last_ID

    def get_summary_results_row(self):
        return self.row_summary_results

    def get_data_row(self):
        return self.row_data

